"""Excel parser for importing data from XLSX files."""

import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger(__name__)


@dataclass
class BOMItem:
    """Bill of Materials item."""

    material: str
    dimension: str
    quantity: float
    unit: str
    note: str


@dataclass
class PriceItem:
    """Price list item."""

    code: str
    name: str
    unit: str
    price_per_unit: float
    currency: str = "CZK"


class ExcelParser:
    """Parses XLSX files for data import."""

    def __init__(self) -> None:
        """Initialize parser."""
        self.logger = logger.bind(component="excel_parser")

    async def parse_bom(self, file_path: str) -> list[BOMItem]:
        """Parse Bill of Materials from XLSX.

        Expected columns: material, dimension, quantity, unit, note

        Args:
            file_path: Path to XLSX file

        Returns:
            List of BOMItem objects

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns are missing or data is invalid
        """
        self.logger.info("parsing_bom", file_path=file_path)

        def _parse() -> list[BOMItem]:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet found")

            items: list[BOMItem] = []
            headers: dict[str, int] = {}

            # Parse header row
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # First row is header
                    for col_idx, cell_value in enumerate(row):
                        if cell_value:
                            headers[str(cell_value).lower().strip()] = col_idx

                    # Validate required columns
                    required = {"material", "dimension", "quantity", "unit", "note"}
                    missing = required - set(headers.keys())
                    if missing:
                        raise ValueError(f"Missing required columns: {missing}")
                    continue

                # Skip empty rows
                if not any(row):
                    continue

                try:
                    items.append(
                        BOMItem(
                            material=str(row[headers["material"]] or "").strip(),
                            dimension=str(row[headers["dimension"]] or "").strip(),
                            quantity=float(row[headers["quantity"]] or 0),
                            unit=str(row[headers["unit"]] or "").strip(),
                            note=str(row[headers["note"]] or "").strip(),
                        )
                    )
                except (ValueError, TypeError, IndexError) as e:
                    self.logger.warning(
                        "skipping_invalid_bom_row",
                        row=row_idx,
                        error=str(e),
                    )
                    continue

            wb.close()
            return items

        # Run in thread executor to avoid blocking
        items = await asyncio.to_thread(_parse)
        self.logger.info("bom_parsed", count=len(items))
        return items

    async def parse_price_list(self, file_path: str) -> list[PriceItem]:
        """Parse price list from XLSX.

        Expected columns: code, name, unit, price_per_unit, currency

        Args:
            file_path: Path to XLSX file

        Returns:
            List of PriceItem objects

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns are missing or data is invalid
        """
        self.logger.info("parsing_price_list", file_path=file_path)

        def _parse() -> list[PriceItem]:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet found")

            items: list[PriceItem] = []
            headers: dict[str, int] = {}

            # Parse header row
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # First row is header
                    for col_idx, cell_value in enumerate(row):
                        if cell_value:
                            headers[str(cell_value).lower().strip()] = col_idx

                    # Validate required columns
                    required = {"code", "name", "unit", "price_per_unit"}
                    missing = required - set(headers.keys())
                    if missing:
                        raise ValueError(f"Missing required columns: {missing}")
                    continue

                # Skip empty rows
                if not any(row):
                    continue

                try:
                    currency = "CZK"
                    if "currency" in headers and row[headers["currency"]]:
                        currency = str(row[headers["currency"]]).strip()

                    items.append(
                        PriceItem(
                            code=str(row[headers["code"]] or "").strip(),
                            name=str(row[headers["name"]] or "").strip(),
                            unit=str(row[headers["unit"]] or "").strip(),
                            price_per_unit=float(row[headers["price_per_unit"]] or 0),
                            currency=currency,
                        )
                    )
                except (ValueError, TypeError, IndexError) as e:
                    self.logger.warning(
                        "skipping_invalid_price_row",
                        row=row_idx,
                        error=str(e),
                    )
                    continue

            wb.close()
            return items

        # Run in thread executor to avoid blocking
        items = await asyncio.to_thread(_parse)
        self.logger.info("price_list_parsed", count=len(items))
        return items

    async def parse_generic(
        self, file_path: str, sheet_name: str | None = None
    ) -> list[dict[str, Any]]:
        """Parse any XLSX into list of dicts (header row as keys).

        Args:
            file_path: Path to XLSX file
            sheet_name: Optional sheet name to parse (default: active sheet)

        Returns:
            List of dictionaries with header row as keys

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If sheet not found or no data
        """
        self.logger.info(
            "parsing_generic", file_path=file_path, sheet_name=sheet_name
        )

        def _parse() -> list[dict[str, Any]]:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(filename=file_path, read_only=True, data_only=True)

            # Get worksheet
            ws: Worksheet | None = None
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet not found: {sheet_name}")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if ws is None:
                raise ValueError("No worksheet found")

            data: list[dict[str, Any]] = []
            headers: list[str] = []

            # Parse rows
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # First row is header
                    headers = [
                        str(cell).strip() if cell is not None else f"Column{i}"
                        for i, cell in enumerate(row, start=1)
                    ]
                    continue

                # Skip empty rows
                if not any(row):
                    continue

                # Create dict from row
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value

                data.append(row_dict)

            wb.close()
            return data

        # Run in thread executor to avoid blocking
        data = await asyncio.to_thread(_parse)
        self.logger.info("generic_parsed", count=len(data))
        return data
