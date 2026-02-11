"""Excel exporter for generating XLSX files."""

import asyncio
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger(__name__)


class ExcelExporter:
    """Exports data to XLSX files."""

    def __init__(self) -> None:
        """Initialize exporter."""
        self.logger = logger.bind(component="excel_exporter")

    def _style_header_row(self, ws: Worksheet, headers: list[str]) -> None:
        """Apply styling to header row.

        Args:
            ws: Worksheet to style
            headers: List of header names
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content.

        Args:
            ws: Worksheet to adjust
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    logger.warning("column_width_calc_failed", column=column_letter)

            # Set column width with min/max bounds
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_currency(self, ws: Worksheet, col_indices: list[int]) -> None:
        """Format currency columns.

        Args:
            ws: Worksheet to format
            col_indices: List of column indices (1-based) to format as currency
        """
        for col_idx in col_indices:
            for row in range(2, ws.max_row + 1):  # Skip header row
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, int | float):
                    cell.number_format = '#,##0.00 "Kč"'

    async def export_orders(self, orders: list[dict[str, Any]], output_path: str) -> str:
        """Export orders to XLSX with formatted table.

        Args:
            orders: List of order dictionaries
            output_path: Path to output XLSX file

        Returns:
            Output file path on success

        Raises:
            ValueError: If orders list is empty
        """
        self.logger.info("exporting_orders", count=len(orders), output_path=output_path)

        if not orders:
            raise ValueError("Orders list is empty")

        def _export() -> str:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet")

            ws.title = "Zakázky"

            # Define columns
            headers = [
                "Číslo",
                "Zákazník",
                "Název",
                "Datum vytvoření",
                "Termín dodání",
                "Celková částka",
                "Stav",
            ]

            # Write header
            ws.append(headers)
            self._style_header_row(ws, headers)

            # Write data rows
            for order in orders:
                row = [
                    order.get("cislo", ""),
                    order.get("zakaznik", ""),
                    order.get("nazev", ""),
                    order.get("datum_vytvoreni", ""),
                    order.get("termin_dodani", ""),
                    order.get("celkova_castka", 0),
                    order.get("stav", ""),
                ]
                ws.append(row)

            # Format currency column (Celková částka is column 6)
            self._format_currency(ws, [6])

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save workbook
            output = Path(output_path)
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            wb.close()

            return output_path

        # Run in thread executor to avoid blocking
        result = await asyncio.to_thread(_export)
        self.logger.info("orders_exported", output_path=result)
        return result

    async def export_calculations(
        self, calculations: list[dict[str, Any]], output_path: str
    ) -> str:
        """Export calculations with items and totals.

        Args:
            calculations: List of calculation dictionaries with items
            output_path: Path to output XLSX file

        Returns:
            Output file path on success

        Raises:
            ValueError: If calculations list is empty
        """
        self.logger.info(
            "exporting_calculations",
            count=len(calculations),
            output_path=output_path,
        )

        if not calculations:
            raise ValueError("Calculations list is empty")

        def _export() -> str:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet")

            ws.title = "Kalkulace"

            # Define columns
            headers = [
                "Kalkulace ID",
                "Zakázka",
                "Datum",
                "Položka",
                "Množství",
                "Jednotka",
                "Cena/jednotku",
                "Celkem",
            ]

            # Write header
            ws.append(headers)
            self._style_header_row(ws, headers)

            # Write data rows
            for calc in calculations:
                calc_id = calc.get("id", "")
                zakazka = calc.get("zakazka", "")
                datum = calc.get("datum", "")
                items = calc.get("items", [])

                if not items:
                    # No items, write basic info
                    row = [calc_id, zakazka, datum, "", "", "", "", 0]
                    ws.append(row)
                else:
                    # Write each item
                    for item in items:
                        row = [
                            calc_id,
                            zakazka,
                            datum,
                            item.get("nazev", ""),
                            item.get("mnozstvi", 0),
                            item.get("jednotka", ""),
                            item.get("cena_za_jednotku", 0),
                            item.get("celkem", 0),
                        ]
                        ws.append(row)

            # Format currency columns (Cena/jednotku=7, Celkem=8)
            self._format_currency(ws, [7, 8])

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save workbook
            output = Path(output_path)
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            wb.close()

            return output_path

        # Run in thread executor to avoid blocking
        result = await asyncio.to_thread(_export)
        self.logger.info("calculations_exported", output_path=result)
        return result

    async def export_generic(
        self,
        data: list[dict[str, Any]],
        output_path: str,
        sheet_name: str = "Data",
    ) -> str:
        """Export generic data list to XLSX with auto-width columns.

        Args:
            data: List of dictionaries to export
            output_path: Path to output XLSX file
            sheet_name: Name of the worksheet (default: "Data")

        Returns:
            Output file path on success

        Raises:
            ValueError: If data list is empty
        """
        self.logger.info(
            "exporting_generic",
            count=len(data),
            output_path=output_path,
            sheet_name=sheet_name,
        )

        if not data:
            raise ValueError("Data list is empty")

        def _export() -> str:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet")

            ws.title = sheet_name

            # Get all unique keys from all dictionaries
            headers = list(dict.fromkeys(key for row in data for key in row.keys()))

            # Write header
            ws.append(headers)
            self._style_header_row(ws, headers)

            # Write data rows
            for row_data in data:
                row = [row_data.get(header) for header in headers]
                ws.append(row)

            # Detect and format currency columns
            currency_cols: list[int] = []
            for col_idx, header in enumerate(headers, start=1):
                header_lower = header.lower()
                if any(
                    keyword in header_lower
                    for keyword in ["cena", "castka", "price", "amount", "celkem", "total"]
                ):
                    currency_cols.append(col_idx)

            if currency_cols:
                self._format_currency(ws, currency_cols)

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Add metadata
            wb.properties.creator = "inferbox"
            wb.properties.created = datetime.now()

            # Save workbook
            output = Path(output_path)
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            wb.close()

            return output_path

        # Run in thread executor to avoid blocking
        result = await asyncio.to_thread(_export)
        self.logger.info("generic_exported", output_path=result)
        return result

    async def export_material_requirements(
        self,
        items: list[dict[str, Any]],
        total_estimated_cost: Decimal | None = None,
        order_count: int = 0,
    ) -> bytes:
        """Export aggregated material requirements to Excel (BOM / nákupní seznam).

        Args:
            items: List of MaterialRequirementItem dictionaries with keys:
                   material_name, material_grade, total_quantity, unit,
                   estimated_unit_price, total_price, order_numbers, supplier
            total_estimated_cost: Total estimated cost across all materials
            order_count: Number of orders included in the report

        Returns:
            Excel file as bytes (.xlsx)
        """
        self.logger.info(
            "exporting_material_requirements",
            count=len(items),
            total_cost=str(total_estimated_cost) if total_estimated_cost else "N/A",
            order_count=order_count,
        )

        def _export() -> bytes:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet")

            ws.title = "Materiálová potřeba"

            # Define columns
            headers = [
                "Materiál",
                "Třída materiálu",
                "Celkové množství",
                "Jednotka",
                "Cena/jednotku (Kč)",
                "Celková cena (Kč)",
                "Zakázky",
                "Dodavatel",
            ]

            # Write header
            ws.append(headers)
            self._style_header_row(ws, headers)

            # Write data rows
            for item in items:
                order_numbers_str = ", ".join(item.get("order_numbers", []))
                row = [
                    item.get("material_name", ""),
                    item.get("material_grade") or "",
                    float(item.get("total_quantity", 0)),
                    item.get("unit", ""),
                    float(item.get("estimated_unit_price") or 0),
                    float(item.get("total_price") or 0),
                    order_numbers_str,
                    item.get("supplier") or "",
                ]
                ws.append(row)

            # Add summary row if total cost is available
            if total_estimated_cost:
                ws.append([])  # Empty row
                summary_row = [
                    f"Celkem ({order_count} zakázek)",
                    "",
                    "",
                    "",
                    "",
                    float(total_estimated_cost),
                    "",
                    "",
                ]
                ws.append(summary_row)

                # Bold summary row
                summary_row_idx = ws.max_row
                for col_idx in range(1, 9):
                    cell = ws.cell(row=summary_row_idx, column=col_idx)
                    cell.font = Font(bold=True)

            # Format currency columns (Cena/jednotku=5, Celková cena=6)
            self._format_currency(ws, [5, 6])

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Add metadata
            wb.properties.creator = "inferbox"
            wb.properties.created = datetime.now()
            wb.properties.title = "Materiálová potřeba (BOM)"

            # Save to BytesIO instead of file
            output = BytesIO()
            wb.save(output)
            wb.close()

            output.seek(0)
            return output.read()

        # Run in thread executor to avoid blocking
        result = await asyncio.to_thread(_export)
        self.logger.info("material_requirements_exported", size_bytes=len(result))
        return result
