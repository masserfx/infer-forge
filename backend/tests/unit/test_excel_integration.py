"""Unit tests for Excel parser and exporter."""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.integrations.excel.exporter import ExcelExporter
from app.integrations.excel.parser import ExcelParser


class TestExcelParser:
    """Tests for ExcelParser class."""

    @pytest.fixture
    def parser(self) -> ExcelParser:
        """Create Excel parser instance."""
        return ExcelParser()

    @pytest.fixture
    def bom_xlsx_file(self, tmp_path: Path) -> str:
        """Create a test BOM XLSX file."""
        file_path = tmp_path / "bom_test.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Write header
        ws.append(["material", "dimension", "quantity", "unit", "note"])

        # Write data rows
        ws.append(["Ocel S235JR", "Ø42x3", 10.5, "m", "Trubka bezešvá"])
        ws.append(["Pleč 11375", "200x100x8", 5.0, "ks", "Příruba DN100"])
        ws.append(["Svařenec", "DN50", 3.0, "ks", "Koleno 90°"])

        wb.save(file_path)
        wb.close()

        return str(file_path)

    @pytest.fixture
    def price_list_xlsx_file(self, tmp_path: Path) -> str:
        """Create a test price list XLSX file."""
        file_path = tmp_path / "price_list_test.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Write header
        ws.append(["code", "name", "unit", "price_per_unit", "currency"])

        # Write data rows
        ws.append(["MAT-001", "Ocel S235JR", "m", 250.50, "CZK"])
        ws.append(["MAT-002", "Pleč 11375", "ks", 1500.00, "CZK"])
        ws.append(["SVC-001", "Svaření TIG", "hod", 800.00, "CZK"])

        wb.save(file_path)
        wb.close()

        return str(file_path)

    @pytest.fixture
    def generic_xlsx_file(self, tmp_path: Path) -> str:
        """Create a generic test XLSX file."""
        file_path = tmp_path / "generic_test.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Write header
        ws.append(["ID", "Name", "Value", "Status"])

        # Write data rows
        ws.append([1, "Item A", 100.5, "Active"])
        ws.append([2, "Item B", 200.75, "Inactive"])
        ws.append([3, "Item C", 300.25, "Pending"])

        wb.save(file_path)
        wb.close()

        return str(file_path)

    @pytest.mark.asyncio
    async def test_parse_bom_success(
        self,
        parser: ExcelParser,
        bom_xlsx_file: str,
    ) -> None:
        """Test parsing valid BOM XLSX file."""
        items = await parser.parse_bom(bom_xlsx_file)

        assert len(items) == 3

        # Check first item
        assert items[0].material == "Ocel S235JR"
        assert items[0].dimension == "Ø42x3"
        assert items[0].quantity == 10.5
        assert items[0].unit == "m"
        assert items[0].note == "Trubka bezešvá"

        # Check second item
        assert items[1].material == "Pleč 11375"
        assert items[1].dimension == "200x100x8"
        assert items[1].quantity == 5.0
        assert items[1].unit == "ks"

        # Check third item
        assert items[2].material == "Svařenec"
        assert items[2].dimension == "DN50"
        assert items[2].quantity == 3.0

    @pytest.mark.asyncio
    async def test_parse_bom_missing_file(
        self,
        parser: ExcelParser,
    ) -> None:
        """Test parsing non-existent BOM file raises error."""
        with pytest.raises(FileNotFoundError, match="File not found"):
            await parser.parse_bom("/nonexistent/file.xlsx")

    @pytest.mark.asyncio
    async def test_parse_bom_missing_columns(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test parsing BOM with missing required columns raises error."""
        file_path = tmp_path / "invalid_bom.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Missing 'note' column
        ws.append(["material", "dimension", "quantity", "unit"])
        ws.append(["Steel", "DN50", 10, "m"])

        wb.save(file_path)
        wb.close()

        with pytest.raises(ValueError, match="Missing required columns"):
            await parser.parse_bom(str(file_path))

    @pytest.mark.asyncio
    async def test_parse_bom_skips_empty_rows(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test parsing BOM skips empty rows."""
        file_path = tmp_path / "bom_with_empty.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws.append(["material", "dimension", "quantity", "unit", "note"])
        ws.append(["Ocel", "DN50", 10, "m", "Valid"])
        ws.append([None, None, None, None, None])  # Empty row
        ws.append(["Pleč", "DN100", 5, "ks", "Also valid"])

        wb.save(file_path)
        wb.close()

        items = await parser.parse_bom(str(file_path))

        assert len(items) == 2
        assert items[0].material == "Ocel"
        assert items[1].material == "Pleč"

    @pytest.mark.asyncio
    async def test_parse_bom_handles_invalid_rows(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test parsing BOM skips rows with invalid data."""
        file_path = tmp_path / "bom_invalid_data.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws.append(["material", "dimension", "quantity", "unit", "note"])
        ws.append(["Valid", "DN50", 10.5, "m", "OK"])
        ws.append(["Invalid", "DN100", "not_a_number", "m", "Bad quantity"])
        ws.append(["Also valid", "DN80", 7.0, "ks", "OK"])

        wb.save(file_path)
        wb.close()

        items = await parser.parse_bom(str(file_path))

        # Should skip the invalid row
        assert len(items) == 2
        assert items[0].material == "Valid"
        assert items[1].material == "Also valid"

    @pytest.mark.asyncio
    async def test_parse_price_list_success(
        self,
        parser: ExcelParser,
        price_list_xlsx_file: str,
    ) -> None:
        """Test parsing valid price list XLSX file."""
        items = await parser.parse_price_list(price_list_xlsx_file)

        assert len(items) == 3

        # Check first item
        assert items[0].code == "MAT-001"
        assert items[0].name == "Ocel S235JR"
        assert items[0].unit == "m"
        assert items[0].price_per_unit == 250.50
        assert items[0].currency == "CZK"

        # Check second item
        assert items[1].code == "MAT-002"
        assert items[1].name == "Pleč 11375"
        assert items[1].price_per_unit == 1500.00

    @pytest.mark.asyncio
    async def test_parse_price_list_missing_currency_defaults_czk(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test price list defaults to CZK when currency column missing."""
        file_path = tmp_path / "price_no_currency.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # No currency column
        ws.append(["code", "name", "unit", "price_per_unit"])
        ws.append(["MAT-001", "Material", "m", 100.0])

        wb.save(file_path)
        wb.close()

        items = await parser.parse_price_list(str(file_path))

        assert len(items) == 1
        assert items[0].currency == "CZK"

    @pytest.mark.asyncio
    async def test_parse_price_list_missing_required_columns(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test price list with missing required columns raises error."""
        file_path = tmp_path / "invalid_price.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Missing 'unit' column
        ws.append(["code", "name", "price_per_unit"])
        ws.append(["MAT-001", "Material", 100.0])

        wb.save(file_path)
        wb.close()

        with pytest.raises(ValueError, match="Missing required columns"):
            await parser.parse_price_list(str(file_path))

    @pytest.mark.asyncio
    async def test_parse_generic_success(
        self,
        parser: ExcelParser,
        generic_xlsx_file: str,
    ) -> None:
        """Test parsing generic XLSX file."""
        data = await parser.parse_generic(generic_xlsx_file)

        assert len(data) == 3

        # Check structure
        assert data[0]["ID"] == 1
        assert data[0]["Name"] == "Item A"
        assert data[0]["Value"] == 100.5
        assert data[0]["Status"] == "Active"

        assert data[1]["Name"] == "Item B"
        assert data[2]["Status"] == "Pending"

    @pytest.mark.asyncio
    async def test_parse_generic_with_sheet_name(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test parsing specific sheet by name."""
        file_path = tmp_path / "multisheet.xlsx"

        wb = Workbook()

        # Create first sheet
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])

        # Create second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y", "Z"])
        ws2.append([10, 20, 30])

        wb.save(file_path)
        wb.close()

        # Parse specific sheet
        data = await parser.parse_generic(str(file_path), sheet_name="Sheet2")

        assert len(data) == 1
        assert data[0]["X"] == 10
        assert data[0]["Y"] == 20
        assert data[0]["Z"] == 30

    @pytest.mark.asyncio
    async def test_parse_generic_sheet_not_found(
        self,
        parser: ExcelParser,
        generic_xlsx_file: str,
    ) -> None:
        """Test parsing non-existent sheet raises error."""
        with pytest.raises(ValueError, match="Sheet not found"):
            await parser.parse_generic(generic_xlsx_file, sheet_name="NonExistent")

    @pytest.mark.asyncio
    async def test_parse_generic_skips_empty_rows(
        self,
        parser: ExcelParser,
        tmp_path: Path,
    ) -> None:
        """Test generic parser skips empty rows."""
        file_path = tmp_path / "generic_empty.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws.append(["Col1", "Col2"])
        ws.append([1, 2])
        ws.append([None, None])  # Empty row
        ws.append([3, 4])

        wb.save(file_path)
        wb.close()

        data = await parser.parse_generic(str(file_path))

        assert len(data) == 2
        assert data[0]["Col1"] == 1
        assert data[1]["Col1"] == 3


class TestExcelExporter:
    """Tests for ExcelExporter class."""

    @pytest.fixture
    def exporter(self) -> ExcelExporter:
        """Create Excel exporter instance."""
        return ExcelExporter()

    @pytest.mark.asyncio
    async def test_export_orders_success(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting orders to XLSX."""
        orders = [
            {
                "cislo": "Z-2024-001",
                "zakaznik": "ACME s.r.o.",
                "nazev": "Svařence DN50",
                "datum_vytvoreni": "2024-01-15",
                "termin_dodani": "2024-02-15",
                "celkova_castka": 45000.50,
                "stav": "Rozpracováno",
            },
            {
                "cislo": "Z-2024-002",
                "zakaznik": "Beta Corp.",
                "nazev": "Konstrukce",
                "datum_vytvoreni": "2024-01-20",
                "termin_dodani": "2024-03-01",
                "celkova_castka": 128000.00,
                "stav": "Nová",
            },
        ]

        output_path = tmp_path / "orders_export.xlsx"
        result = await exporter.export_orders(orders, str(output_path))

        assert result == str(output_path)
        assert output_path.exists()

        # Verify content
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None
        assert ws.title == "Zakázky"

        # Check header row
        headers = [cell.value for cell in ws[1]]
        assert "Číslo" in headers
        assert "Zákazník" in headers
        assert "Celková částka" in headers

        # Check data rows
        assert ws.cell(row=2, column=1).value == "Z-2024-001"
        assert ws.cell(row=2, column=2).value == "ACME s.r.o."
        assert ws.cell(row=3, column=1).value == "Z-2024-002"

        wb.close()

    @pytest.mark.asyncio
    async def test_export_orders_empty_list_raises_error(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting empty orders list raises error."""
        output_path = tmp_path / "empty_orders.xlsx"

        with pytest.raises(ValueError, match="Orders list is empty"):
            await exporter.export_orders([], str(output_path))

    @pytest.mark.asyncio
    async def test_export_calculations_success(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting calculations to XLSX."""
        calculations = [
            {
                "id": "CALC-001",
                "zakazka": "Z-2024-001",
                "datum": "2024-01-15",
                "items": [
                    {
                        "nazev": "Trubka DN50",
                        "mnozstvi": 10.5,
                        "jednotka": "m",
                        "cena_za_jednotku": 250.00,
                        "celkem": 2625.00,
                    },
                    {
                        "nazev": "Svařování",
                        "mnozstvi": 5.0,
                        "jednotka": "hod",
                        "cena_za_jednotku": 800.00,
                        "celkem": 4000.00,
                    },
                ],
            },
        ]

        output_path = tmp_path / "calc_export.xlsx"
        result = await exporter.export_calculations(calculations, str(output_path))

        assert result == str(output_path)
        assert output_path.exists()

        # Verify content
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None
        assert ws.title == "Kalkulace"

        # Check header
        headers = [cell.value for cell in ws[1]]
        assert "Kalkulace ID" in headers
        assert "Položka" in headers
        assert "Množství" in headers

        # Check data (2 items from calculation)
        assert ws.cell(row=2, column=1).value == "CALC-001"
        assert ws.cell(row=2, column=4).value == "Trubka DN50"
        assert ws.cell(row=3, column=4).value == "Svařování"

        wb.close()

    @pytest.mark.asyncio
    async def test_export_calculations_empty_items(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting calculation with empty items list."""
        calculations = [
            {
                "id": "CALC-001",
                "zakazka": "Z-2024-001",
                "datum": "2024-01-15",
                "items": [],
            },
        ]

        output_path = tmp_path / "calc_empty_items.xlsx"
        await exporter.export_calculations(calculations, str(output_path))

        assert output_path.exists()

        # Verify basic row was written
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=2, column=1).value == "CALC-001"

        wb.close()

    @pytest.mark.asyncio
    async def test_export_generic_success(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting generic data to XLSX."""
        data = [
            {"ID": 1, "Name": "Item A", "Price": 100.50, "Status": "Active"},
            {"ID": 2, "Name": "Item B", "Price": 200.75, "Status": "Inactive"},
            {"ID": 3, "Name": "Item C", "Price": 300.25, "Status": "Pending"},
        ]

        output_path = tmp_path / "generic_export.xlsx"
        result = await exporter.export_generic(
            data,
            str(output_path),
            sheet_name="TestData",
        )

        assert result == str(output_path)
        assert output_path.exists()

        # Verify content
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None
        assert ws.title == "TestData"

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert headers == ["ID", "Name", "Price", "Status"]

        # Check data
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "Item A"
        assert ws.cell(row=3, column=1).value == 2

        # Check metadata
        assert wb.properties.creator == "inferbox"

        wb.close()

    @pytest.mark.asyncio
    async def test_export_generic_empty_data_raises_error(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporting empty data list raises error."""
        output_path = tmp_path / "empty_generic.xlsx"

        with pytest.raises(ValueError, match="Data list is empty"):
            await exporter.export_generic([], str(output_path))

    @pytest.mark.asyncio
    async def test_export_generic_creates_parent_dirs(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporter creates parent directories if needed."""
        data = [{"A": 1, "B": 2}]

        # Path with non-existent parent directories
        output_path = tmp_path / "subdir1" / "subdir2" / "export.xlsx"

        await exporter.export_generic(data, str(output_path))

        assert output_path.exists()
        assert output_path.parent.exists()

    @pytest.mark.asyncio
    async def test_export_generic_handles_mixed_keys(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test generic export handles dictionaries with different keys."""
        data = [
            {"A": 1, "B": 2, "C": 3},
            {"A": 4, "B": 5},  # Missing 'C'
            {"A": 6, "C": 7, "D": 8},  # Missing 'B', extra 'D'
        ]

        output_path = tmp_path / "mixed_keys.xlsx"
        await exporter.export_generic(data, str(output_path))

        assert output_path.exists()

        # Verify all keys are in headers
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None

        headers = [cell.value for cell in ws[1]]
        assert "A" in headers
        assert "B" in headers
        assert "C" in headers
        assert "D" in headers

        wb.close()

    @pytest.mark.asyncio
    async def test_export_detects_currency_columns(
        self,
        exporter: ExcelExporter,
        tmp_path: Path,
    ) -> None:
        """Test exporter auto-detects and formats currency columns."""
        data = [
            {"ID": 1, "cena": 1000.50, "celkem": 5000.00},
            {"ID": 2, "cena": 2000.75, "celkem": 10000.00},
        ]

        output_path = tmp_path / "currency_test.xlsx"
        await exporter.export_generic(data, str(output_path))

        # Load and check number format
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws is not None

        # Currency columns should have special format
        # (Column 2: cena, Column 3: celkem)
        price_cell = ws.cell(row=2, column=2)
        assert 'Kč' in price_cell.number_format

        wb.close()
