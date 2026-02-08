"""Tests for material requirements reporting and export."""

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.integrations.excel.exporter import ExcelExporter
from app.models import (
    Calculation,
    CalculationItem,
    CalculationStatus,
    CostType,
    Customer,
    MaterialPrice,
    Order,
    OrderStatus,
)
from app.services.reporting import ReportingService


@pytest.fixture
async def test_customer(test_db: AsyncSession) -> Customer:
    """Create test customer."""
    customer = Customer(
        company_name="Test Company s.r.o.",
        ico="12345678",
        contact_name="Jan Novák",
        email="test@example.com",
    )
    test_db.add(customer)
    await test_db.commit()
    await test_db.refresh(customer)
    return customer


@pytest.fixture
async def test_order(test_db: AsyncSession, test_customer: Customer) -> Order:
    """Create test order."""
    order = Order(
        number="TEST-001",
        customer_id=test_customer.id,
        status=OrderStatus.OBJEDNAVKA,
        due_date=date.today() + timedelta(days=30),
    )
    test_db.add(order)
    await test_db.commit()
    await test_db.refresh(order)
    return order


@pytest.fixture
async def test_calculation(test_db: AsyncSession, test_order: Order) -> Calculation:
    """Create test calculation."""
    calc = Calculation(
        order_id=test_order.id,
        name="Test Calculation",
        status=CalculationStatus.APPROVED,
        material_total=Decimal("10000.00"),
        labor_total=Decimal("5000.00"),
        margin_percent=Decimal("15.00"),
        total_price=Decimal("17250.00"),
    )
    test_db.add(calc)
    await test_db.commit()
    await test_db.refresh(calc)
    return calc


@pytest.fixture
async def test_material_items(
    test_db: AsyncSession, test_calculation: Calculation
) -> list[CalculationItem]:
    """Create test material items."""
    items = [
        CalculationItem(
            calculation_id=test_calculation.id,
            cost_type=CostType.MATERIAL,
            name="Ocel S235JR",
            description="Plech 10mm",
            quantity=Decimal("100.5"),
            unit="kg",
            unit_price=Decimal("25.50"),
            total_price=Decimal("2562.75"),
        ),
        CalculationItem(
            calculation_id=test_calculation.id,
            cost_type=CostType.MATERIAL,
            name="Trubka DN100",
            description="Svařovaná trubka",
            quantity=Decimal("50.0"),
            unit="m",
            unit_price=Decimal("150.00"),
            total_price=Decimal("7500.00"),
        ),
        CalculationItem(
            calculation_id=test_calculation.id,
            cost_type=CostType.LABOR,
            name="Svařování",
            quantity=Decimal("10.0"),
            unit="hod",
            unit_price=Decimal("500.00"),
            total_price=Decimal("5000.00"),
        ),
    ]
    for item in items:
        test_db.add(item)
    await test_db.commit()
    for item in items:
        await test_db.refresh(item)
    return items


@pytest.fixture
async def test_material_prices(test_db: AsyncSession) -> list[MaterialPrice]:
    """Create test material prices."""
    today = date.today()
    prices = [
        MaterialPrice(
            name="Ocel S235JR",
            material_grade="S235JR",
            form="plech",
            dimension="10mm",
            unit="kg",
            unit_price=Decimal("26.00"),
            supplier="Ferona",
            valid_from=today - timedelta(days=30),
            valid_to=None,
            is_active=True,
        ),
        MaterialPrice(
            name="Trubka DN100",
            material_grade="P235GH",
            form="trubka",
            dimension="DN100",
            unit="m",
            unit_price=Decimal("155.00"),
            supplier="ArcelorMittal",
            valid_from=today - timedelta(days=60),
            valid_to=None,
            is_active=True,
        ),
    ]
    for price in prices:
        test_db.add(price)
    await test_db.commit()
    for price in prices:
        await test_db.refresh(price)
    return prices


@pytest.mark.asyncio
async def test_material_requirements_basic(
    test_db: AsyncSession,
    test_order: Order,
    test_calculation: Calculation,
    test_material_items: list[CalculationItem],
    test_material_prices: list[MaterialPrice],
) -> None:
    """Test basic material requirements extraction."""
    service = ReportingService(test_db)

    result = await service.get_material_requirements(order_ids=[test_order.id])

    assert result.order_count == 1
    assert len(result.items) == 2  # Only material items, not labor

    # Check Ocel S235JR
    steel_item = next((item for item in result.items if "Ocel" in item.material_name), None)
    assert steel_item is not None
    assert steel_item.material_grade == "S235JR"
    assert steel_item.total_quantity == Decimal("100.5")
    assert steel_item.unit == "kg"
    assert steel_item.estimated_unit_price == Decimal("26.00")  # From MaterialPrice
    assert steel_item.supplier == "Ferona"
    assert test_order.number in steel_item.order_numbers

    # Check Trubka DN100
    pipe_item = next((item for item in result.items if "Trubka" in item.material_name), None)
    assert pipe_item is not None
    assert pipe_item.material_grade == "P235GH"
    assert pipe_item.total_quantity == Decimal("50.0")
    assert pipe_item.unit == "m"
    assert pipe_item.estimated_unit_price == Decimal("155.00")  # From MaterialPrice
    assert pipe_item.supplier == "ArcelorMittal"

    # Check total cost
    assert result.total_estimated_cost is not None
    expected_cost = (Decimal("100.5") * Decimal("26.00")) + (
        Decimal("50.0") * Decimal("155.00")
    )
    assert result.total_estimated_cost == expected_cost


@pytest.mark.asyncio
async def test_material_requirements_aggregation(
    test_db: AsyncSession, test_customer: Customer
) -> None:
    """Test material aggregation across multiple orders."""
    # Create two orders
    order1 = Order(
        number="ORDER-001",
        customer_id=test_customer.id,
        status=OrderStatus.VYROBA,
    )
    order2 = Order(
        number="ORDER-002",
        customer_id=test_customer.id,
        status=OrderStatus.VYROBA,
    )
    test_db.add(order1)
    test_db.add(order2)
    await test_db.commit()
    await test_db.refresh(order1)
    await test_db.refresh(order2)

    # Create calculations for both orders
    calc1 = Calculation(
        order_id=order1.id,
        name="Calc 1",
        status=CalculationStatus.APPROVED,
    )
    calc2 = Calculation(
        order_id=order2.id,
        name="Calc 2",
        status=CalculationStatus.APPROVED,
    )
    test_db.add(calc1)
    test_db.add(calc2)
    await test_db.commit()
    await test_db.refresh(calc1)
    await test_db.refresh(calc2)

    # Add same material to both calculations
    item1 = CalculationItem(
        calculation_id=calc1.id,
        cost_type=CostType.MATERIAL,
        name="Ocel S235JR",
        quantity=Decimal("100.0"),
        unit="kg",
        unit_price=Decimal("25.00"),
        total_price=Decimal("2500.00"),
    )
    item2 = CalculationItem(
        calculation_id=calc2.id,
        cost_type=CostType.MATERIAL,
        name="Ocel S235JR",  # Same material
        quantity=Decimal("50.0"),
        unit="kg",
        unit_price=Decimal("26.00"),
        total_price=Decimal("1300.00"),
    )
    test_db.add(item1)
    test_db.add(item2)
    await test_db.commit()

    service = ReportingService(test_db)
    result = await service.get_material_requirements(order_ids=[order1.id, order2.id])

    assert result.order_count == 2
    assert len(result.items) == 1  # Aggregated to single material

    steel_item = result.items[0]
    assert steel_item.material_name == "Ocel S235JR"
    assert steel_item.total_quantity == Decimal("150.0")  # 100 + 50
    assert steel_item.unit == "kg"
    # Price should be average of the two
    assert steel_item.estimated_unit_price == Decimal("25.50")  # (25 + 26) / 2
    assert len(steel_item.order_numbers) == 2
    assert "ORDER-001" in steel_item.order_numbers
    assert "ORDER-002" in steel_item.order_numbers


@pytest.mark.asyncio
async def test_material_requirements_empty(test_db: AsyncSession, test_customer: Customer) -> None:
    """Test material requirements with no calculations."""
    order = Order(
        number="EMPTY-001",
        customer_id=test_customer.id,
        status=OrderStatus.POPTAVKA,
    )
    test_db.add(order)
    await test_db.commit()
    await test_db.refresh(order)

    service = ReportingService(test_db)
    result = await service.get_material_requirements(order_ids=[order.id])

    assert result.order_count == 1
    assert len(result.items) == 0
    assert result.total_estimated_cost == Decimal("0")


@pytest.mark.asyncio
async def test_material_requirements_status_filter(
    test_db: AsyncSession, test_customer: Customer
) -> None:
    """Test filtering by order status."""
    # Create orders with different statuses
    order_poptavka = Order(
        number="POPT-001",
        customer_id=test_customer.id,
        status=OrderStatus.POPTAVKA,
    )
    order_vyroba = Order(
        number="VYR-001",
        customer_id=test_customer.id,
        status=OrderStatus.VYROBA,
    )
    test_db.add(order_poptavka)
    test_db.add(order_vyroba)
    await test_db.commit()
    await test_db.refresh(order_poptavka)
    await test_db.refresh(order_vyroba)

    # Add calculations to both
    calc_poptavka = Calculation(
        order_id=order_poptavka.id,
        name="Calc Poptavka",
        status=CalculationStatus.DRAFT,
    )
    calc_vyroba = Calculation(
        order_id=order_vyroba.id,
        name="Calc Vyroba",
        status=CalculationStatus.APPROVED,
    )
    test_db.add(calc_poptavka)
    test_db.add(calc_vyroba)
    await test_db.commit()
    await test_db.refresh(calc_poptavka)
    await test_db.refresh(calc_vyroba)

    # Add material items
    item_poptavka = CalculationItem(
        calculation_id=calc_poptavka.id,
        cost_type=CostType.MATERIAL,
        name="Material A",
        quantity=Decimal("10.0"),
        unit="kg",
        unit_price=Decimal("100.00"),
        total_price=Decimal("1000.00"),
    )
    item_vyroba = CalculationItem(
        calculation_id=calc_vyroba.id,
        cost_type=CostType.MATERIAL,
        name="Material B",
        quantity=Decimal("20.0"),
        unit="kg",
        unit_price=Decimal("200.00"),
        total_price=Decimal("4000.00"),
    )
    test_db.add(item_poptavka)
    test_db.add(item_vyroba)
    await test_db.commit()

    service = ReportingService(test_db)

    # Filter only VYROBA status
    result = await service.get_material_requirements(status_filter=["vyroba"])

    assert result.order_count == 1
    assert len(result.items) == 1
    assert result.items[0].material_name == "Material B"
    assert "VYR-001" in result.items[0].order_numbers


@pytest.mark.asyncio
async def test_material_requirements_with_best_price(
    test_db: AsyncSession, test_order: Order, test_calculation: Calculation
) -> None:
    """Test material price matching with MaterialPrice database."""
    today = date.today()

    # Add material item WITHOUT unit price
    item = CalculationItem(
        calculation_id=test_calculation.id,
        cost_type=CostType.MATERIAL,
        name="Ocel S355J2",
        quantity=Decimal("200.0"),
        unit="kg",
        unit_price=Decimal("0.00"),  # No price in calculation
        total_price=Decimal("0.00"),
    )
    test_db.add(item)

    # Add MaterialPrice for this material
    price = MaterialPrice(
        name="Ocel S355J2",
        material_grade="S355J2",
        form="plech",
        unit="kg",
        unit_price=Decimal("32.50"),
        supplier="Test Supplier",
        valid_from=today - timedelta(days=10),
        valid_to=None,
        is_active=True,
    )
    test_db.add(price)
    await test_db.commit()

    service = ReportingService(test_db)
    result = await service.get_material_requirements(order_ids=[test_order.id])

    # Find the steel item
    steel_item = next(
        (item for item in result.items if "S355J2" in item.material_name), None
    )
    assert steel_item is not None
    assert steel_item.estimated_unit_price == Decimal("32.50")  # From MaterialPrice
    assert steel_item.supplier == "Test Supplier"
    assert steel_item.total_price == Decimal("200.0") * Decimal("32.50")


@pytest.mark.asyncio
async def test_material_requirements_excel_export(
    test_db: AsyncSession,
    test_order: Order,
    test_calculation: Calculation,
    test_material_items: list[CalculationItem],
) -> None:
    """Test Excel export of material requirements."""
    service = ReportingService(test_db)
    result = await service.get_material_requirements(order_ids=[test_order.id])

    # Convert to dict for export
    items_dict = [item.model_dump() for item in result.items]

    exporter = ExcelExporter()
    excel_bytes = await exporter.export_material_requirements(
        items=items_dict,
        total_estimated_cost=result.total_estimated_cost,
        order_count=result.order_count,
    )

    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # Verify Excel contents
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws is not None
    assert ws.title == "Materiálová potřeba"

    # Check headers
    headers = [cell.value for cell in ws[1]]
    assert "Materiál" in headers
    assert "Třída materiálu" in headers
    assert "Celkové množství" in headers
    assert "Jednotka" in headers
    assert "Cena/jednotku (Kč)" in headers
    assert "Celková cena (Kč)" in headers
    assert "Zakázky" in headers
    assert "Dodavatel" in headers

    # Check that we have data rows (2 material items + header + summary = 4 rows minimum)
    assert ws.max_row >= 4

    # Check for summary row
    last_row_values = [cell.value for cell in ws[ws.max_row]]
    assert any("Celkem" in str(val) for val in last_row_values if val is not None)


@pytest.mark.asyncio
async def test_material_requirements_no_orders(test_db: AsyncSession) -> None:
    """Test material requirements when no orders match filter."""
    service = ReportingService(test_db)

    # Query with non-existent order ID
    result = await service.get_material_requirements(order_ids=[uuid4()])

    assert result.order_count == 0
    assert len(result.items) == 0
    # When no orders found, service returns cost=0 (not None)
    assert result.total_estimated_cost == Decimal("0")


@pytest.mark.asyncio
async def test_material_requirements_default_active_orders(
    test_db: AsyncSession, test_customer: Customer
) -> None:
    """Test default behavior (OBJEDNAVKA, VYROBA) when no filters provided."""
    # Create orders with different statuses
    order_objednavka = Order(
        number="OBJ-001",
        customer_id=test_customer.id,
        status=OrderStatus.OBJEDNAVKA,
    )
    order_vyroba = Order(
        number="VYR-001",
        customer_id=test_customer.id,
        status=OrderStatus.VYROBA,
    )
    order_dokonceno = Order(
        number="DOK-001",
        customer_id=test_customer.id,
        status=OrderStatus.DOKONCENO,
    )
    test_db.add(order_objednavka)
    test_db.add(order_vyroba)
    test_db.add(order_dokonceno)
    await test_db.commit()
    await test_db.refresh(order_objednavka)
    await test_db.refresh(order_vyroba)
    await test_db.refresh(order_dokonceno)

    # Add calculations
    calc_obj = Calculation(
        order_id=order_objednavka.id, name="Calc OBJ", status=CalculationStatus.APPROVED
    )
    calc_vyr = Calculation(
        order_id=order_vyroba.id, name="Calc VYR", status=CalculationStatus.APPROVED
    )
    calc_dok = Calculation(
        order_id=order_dokonceno.id, name="Calc DOK", status=CalculationStatus.APPROVED
    )
    test_db.add(calc_obj)
    test_db.add(calc_vyr)
    test_db.add(calc_dok)
    await test_db.commit()
    await test_db.refresh(calc_obj)
    await test_db.refresh(calc_vyr)
    await test_db.refresh(calc_dok)

    # Add material items
    item_obj = CalculationItem(
        calculation_id=calc_obj.id,
        cost_type=CostType.MATERIAL,
        name="Material OBJ",
        quantity=Decimal("10.0"),
        unit="kg",
        unit_price=Decimal("10.00"),
        total_price=Decimal("100.00"),
    )
    item_vyr = CalculationItem(
        calculation_id=calc_vyr.id,
        cost_type=CostType.MATERIAL,
        name="Material VYR",
        quantity=Decimal("20.0"),
        unit="kg",
        unit_price=Decimal("20.00"),
        total_price=Decimal("400.00"),
    )
    item_dok = CalculationItem(
        calculation_id=calc_dok.id,
        cost_type=CostType.MATERIAL,
        name="Material DOK",
        quantity=Decimal("30.0"),
        unit="kg",
        unit_price=Decimal("30.00"),
        total_price=Decimal("900.00"),
    )
    test_db.add(item_obj)
    test_db.add(item_vyr)
    test_db.add(item_dok)
    await test_db.commit()

    service = ReportingService(test_db)

    # Call without any filters - should only get OBJEDNAVKA and VYROBA
    result = await service.get_material_requirements()

    assert result.order_count == 2  # Only OBJ and VYR, not DOK
    assert len(result.items) == 2

    material_names = {item.material_name for item in result.items}
    assert "Material OBJ" in material_names
    assert "Material VYR" in material_names
    assert "Material DOK" not in material_names
